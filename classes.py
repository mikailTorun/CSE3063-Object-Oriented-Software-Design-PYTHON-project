import xlrd
from openpyxl.styles import Font, Alignment
import csv
import glob
from tkinter import *
from tkinter import filedialog
import tkinter as tk
from openpyxl import Workbook
import re

from tkinter import filedialog, messagebox, ttk
import pandas as pd


class Student:
    def __init__(self, no, id_number, name, surname):
        self.__no = no
        self.__id = id_number
        self.__name = name
        self.__surname = surname
        self.__polls = []

    def get_no(self):
        return self.__no

    def get_id(self):
        return self.__id

    def get_name(self):
        return self.__name

    def get_surname(self):
        return self.__surname

    def get_polls(self):
        return self.__polls


class StudentList:
    def __init__(self, file_name):
        self.__students = []
        self.__sheet = xlrd.open_workbook(file_name).sheet_by_index(0)
        self.__generate_students()

    def __generate_students(self):
        no = 1
        for i in range(self.__sheet.col(0).__len__()):
            row = self.__sheet.row(i)
            if re.match(r'^-?\d+(?:\.\d+)?$', str(row.__getitem__(1).value)):
                self.__students.append(Student(no, row.__getitem__(2).value,
                                               row.__getitem__(4).value, row.__getitem__(7).value))
                no += 1

    def get_student(self, name_surname):
        searched = None
        similarity = 0.0
        name_surname = str(name_surname).replace('İ', 'i').replace('I', 'ı').lower().replace(" ", "")
        for student in self.__students:
            s = (student.get_name() + student.get_surname()).replace('İ', 'i').replace('I', 'ı')
            s = s.lower().replace(" ", "")
            max_length = max(name_surname.__len__(), s.__len__())
            if name_surname.__len__() >= s.__len__():
                min_name = s
                max_name = name_surname
            else:
                min_name = name_surname
                max_name = s
            counter = 0
            for i in range(min_name.__len__() - 2):
                if max_name.__contains__(min_name[i] + min_name[i + 1] + min_name[i + 2]):
                    counter += 1
            if counter / max_length > similarity:
                similarity = counter / max_length
                searched = student
        return searched

    def get_students(self):
        return self.__students


class Results:
    def __init__(self, student_list):
        self.__book = Workbook()
        self.__student_list = student_list
        self.__sheet_num = 0

    def __write_students(self, sheet):
        self.column_title(sheet, 'A', "NO", 5)
        self.column_title(sheet, 'B', "ID", 12)
        self.column_title(sheet, 'C', "Name", 20)
        self.column_title(sheet, 'D', "Surname", 20)
        no = 2
        for student in self.__student_list:
            self.add_cell(sheet, 'A', no, student.get_no())
            self.add_cell(sheet, 'B', no, student.get_id())
            self.add_cell(sheet, 'C', no, student.get_name())
            self.add_cell(sheet, 'D', no, student.get_surname())
            no += 1

    def new_sheet(self, sheet_name):
        if self.__sheet_num == 0:
            sheet = self.__book.active
            self.__sheet_num += 1
        else:
            sheet = self.__book.create_sheet(sheet_name, self.__sheet_num)
            self.__sheet_num += 1
        sheet.title = sheet_name
        self.__write_students(sheet)

    def column_title(self, sheet, column, column_title, column_width):
        sheet[column + '1'].font = Font(bold=True)
        sheet[column + '1'].alignment = Alignment(horizontal='left')
        sheet[column + '1'] = column_title
        sheet.column_dimensions[column].width = column_width

    def add_cell(self, sheet, column, row, data):
        sheet[column + str(row)].alignment = Alignment(horizontal='left')
        sheet[column + str(row)] = data

    def save_book(self):
        self.__book.save("results.xlsx")

    def get_book(self):
        return self.__book


class Question:
    def __init__(self, question_text, answer_text):
        self.__question_text = question_text
        self.__answer_text = answer_text

    def get_question(self):
        return self.__question_text

    def get_answer(self):
        return self.__answer_text


class Poll:
    def __init__(self, questions):
        self.__questions = questions
        self.__poll_name = ""
        self.__poll_date = ""

    def get_questions(self):
        return self.__questions

    def get_poll_name(self):
        return self.__poll_name

    def set_poll_name(self, poll_name):
        self.__poll_name = poll_name

    def set_poll_date(self, poll_date):
        self.__poll_date = poll_date

    def get_poll_date(self):
        return self.__poll_date


class Reports:
    def __init__(self, file_path, student_list):
        self.__file_path = file_path
        self.__student_list = student_list

    def read_reports(self):
        files = glob.glob(self.__file_path + "/*.csv")
        for file in files:
            with open(file, encoding="utf8") as file_name:
                reader = csv.reader(file_name)
                for line in reader:
                    if line.__getitem__(0).isdigit() is not True:
                        continue
                    student = self.__student_list.get_student(line[1])
                    questions = []
                    i = 4
                    while i < line.__len__():
                        if line[i].replace(" ", "") != "":
                            questions.append(Question(line[i], line[i + 1]))
                        i += 2
                    poll = Poll(questions)
                    poll.set_poll_date(line[3])
                    student.get_polls().append(poll)


class Attendance:
    def __init__(self, student_list, results):
        self.__student_list = student_list
        self.__results = results
        self.__results.new_sheet("Attendance")
        self.__date_list = []
        results.column_title(results.get_book()["Attendance"], 'E', 'Number of Attendance', 25)
        results.column_title(results.get_book()["Attendance"], 'F', 'Attendance Rate', 25)
        results.column_title(results.get_book()["Attendance"], 'G', 'Attendance Percentage', 25)

    def __get_total_lessons_number(self):
        total_lessons_number = 0
        for student in self.__student_list:
            self.__date_list = []
            lessons_number = 0
            for poll in student.get_polls():
                if self.is_new_attend(poll):
                    lessons_number += 1
                    if lessons_number > total_lessons_number:
                        total_lessons_number = lessons_number
        return total_lessons_number

    def add_attendance(self):
        total_lessons_number = self.__get_total_lessons_number()
        row = 2
        for student in self.__student_list:
            self.__date_list = []
            attendance_number = 0
            for poll in student.get_polls():
                if self.is_new_attend(poll):
                    attendance_number += 1
            self.__results.add_cell(self.__results.get_book()["Attendance"], 'E', row, attendance_number)
            self.__results.add_cell(self.__results.get_book()["Attendance"], 'F', row,
                                    float(attendance_number) / total_lessons_number)
            self.__results.add_cell(self.__results.get_book()["Attendance"], 'G', row,
                                    str(int((float(attendance_number) / total_lessons_number) * 100)) + " %")
            row += 1

    def is_new_attend(self, poll):
        date_student = poll.get_poll_date().split()
        month_student = date_student[0]
        day_student = date_student[1]
        hour_student = date_student[3].split(":")[0]
        for date in self.__date_list:
            date_list = date.split()
            month_list = date_list[0]
            day_list = date_list[1]
            hour_list = date_list[3].split(":")[0]
            if month_student == month_list and day_student == day_list and hour_student == hour_list:
                return False
        self.__date_list.append(poll.get_poll_date())
        return True


class AnswerKeys:
    def __init__(self, file_path):
        self.__file_path = file_path

    def get_answer_keys(self):
        files = glob.glob(self.__file_path + "/*.csv")
        answer_keys = []
        questions = None
        for file in files:
            with open(file, encoding="utf8") as file_name:
                reader = csv.reader(file_name)
                for line in reader:
                    if line.__len__() == 1:
                        questions = []
                        poll = Poll(questions)
                        poll.set_poll_name(line[0])
                        answer_keys.append(poll)
                    else:
                        questions.append(Question(line[0], line[1]))
        return answer_keys


class CheckAnswers:
    def __init__(self, student_list, results, answer_keys):
        self.__student_list = student_list
        self.__results = results
        self.__answer_keys = answer_keys

    def check(self):
        results = self.__results
        book = results.get_book()
        for answer_key in self.__answer_keys:
            answer_name = answer_key.get_poll_name()
            self.add_sheet(answer_key.get_poll_name(), len(answer_key.get_questions()))
            results.column_title(book[answer_name], 'O', 'Success Rate', 15)
            results.column_title(book[answer_name], 'P', 'Success Percentage', 20)
            for student in self.__student_list:
                for poll in student.get_polls():
                    if self.poll_control(answer_key.get_questions(), poll.get_questions()):
                        chr_number = ord('D')
                        true_number = 0
                        for true_answer, student_answer in zip(answer_key.get_questions(), poll.get_questions()):
                            chr_number += 1
                            true_number += int(true_answer.get_answer() == student_answer.get_answer())
                            results.add_cell(book[answer_name], chr(chr_number), student.get_no() + 1,
                                             int(true_answer.get_answer() == student_answer.get_answer()))
                        success_rate = float(true_number) / len(answer_key.get_questions())
                        results.add_cell(book[answer_name], 'O', student.get_no() + 1, success_rate)
                        results.add_cell(book[answer_name], 'P', student.get_no() + 1,
                                         str(int(success_rate * 100)) + '%')

    def add_sheet(self, poll_name, questions_number):
        self.__results.new_sheet(poll_name)
        chr_number = ord('D')
        for i in range(questions_number):
            chr_number += 1
            self.__results.column_title(self.__results.get_book()[poll_name], chr(chr_number),
                                        'Q' + str(chr_number - 68), 5)

    def poll_control(self, questions_a, questions_b):
        if len(questions_a) != len(questions_b):
            return False
        for question_a, question_b in zip(questions_a, questions_b):
            if question_a.get_question() != question_b.get_question():
                return False
        return True


class GeneralResults:
    def __init__(self, student_list, results):
        self.__student_list = student_list
        self.__results = results
        self.__number_of_pages = self.get_number_of_pages()

    def get_number_of_pages(self):
        number_of_pages = 0
        for sheet in self.__results.get_book():
            if "Poll" in sheet.title:
                number_of_pages += 1
        return number_of_pages

    def print_result(self):
        self.__results.new_sheet("general_stats")
        self.__results.column_title(self.__results.get_book()["general_stats"], "E", "General Success Rate", 20)
        self.__results.column_title(self.__results.get_book()["general_stats"], "F", "General Success Percentage", 25)
        total = 0.0
        i = 2
        while i < len(self.__student_list.get_students()) + 2:
            for sheet in self.__results.get_book():
                if "Poll" in sheet.title:
                    cell = sheet.cell(i, 15)
                    if cell.value is None:
                        cell.value=0
                        cell.alignment = Alignment(horizontal='left')
                    total = total + cell.value
                self.__results.add_cell(self.__results.get_book()['general_stats'], "E", i,
                                        total / self.__number_of_pages)
                self.__results.add_cell(self.__results.get_book()['general_stats'], "F", i,
                                        str(int((total / self.__number_of_pages) * 100)) + ' %')
            total = 0.0
            i += 1


class GUI:
    def __init__(self):
        self.__window = tk.Tk()
        self.__file = ""
        self.__window.geometry("1500x725")
        self.__window.configure(bg='SkyBlue3')
        self.__window.title("Zoom Attendance and Poll Report")
        self.__student_list_file = ""
        self.__reports_file = ""
        self.__answers_file = ""
        Label(self.__window, text='Zoom Attendance and Poll Report',
              font=('Verdana', 10), bg='SkyBlue3').pack(side=TOP, pady=10)
        Label(self.__window, text='Group 19', font=('Verdana', 8), bg='SkyBlue3').pack(side=BOTTOM, pady=10)
        self.tv1 = ""

    def get_file(self):
        return self.__file

    def student_list(self):
        student_list_file = filedialog.askopenfilename(title="select a student list")
        self.__student_list_file = student_list_file

    def reports(self):
        reports_file = filedialog.askdirectory(title="select a reports")
        self.__reports_file = reports_file

    def answers(self):
        answers_file = filedialog.askdirectory(title="select a answers file")
        self.__answers_file = answers_file

    def config_button_xls(self, button):
        button.configure(
            width=25,
            height=2,
            bg="turquoise1",
            fg="black",
            font=('Verdana', 20)
        )
        button.pack()

    def config_button_folder(self, button):
        button.configure(
            width=25,
            height=2,
            bg="turquoise3",
            fg="black",
            font=('Verdana', 20)
        )
        button.pack()

    def config_buttons_start(self, button):
        button.configure(
            width=25,
            height=2,
            bg="cyan3",
            fg="black",
            font=('Verdana', 20)
        )
        button.pack()

    def start_process(self):
        student_list = StudentList(self.__student_list_file)
        Reports(self.__reports_file, student_list).read_reports()
        results = Results(student_list.get_students())
        Attendance(student_list.get_students(), results).add_attendance()
        answer_keys = AnswerKeys(self.__answers_file).get_answer_keys()
        CheckAnswers(student_list.get_students(), results, answer_keys).check()
        GeneralResults(student_list,results).print_result()
        results.save_book()
        conclusion = tk.Label(text="Process Completed", font=('Verdana', 7), bg='SkyBlue3')
        conclusion.pack()

    def compile_window(self):
        self.__window.mainloop()

    def display_result_file(self):
        root = tk.Tk()

        root.geometry("1700x700")
        root.pack_propagate(False)
        root.resizable(0, 0)

        # Frame for TreeView
        frame1 = tk.LabelFrame(root, text="Excel Data")
        frame1.place(height=400, width=800)

        # Frame for open file dialog
        file_frame = tk.LabelFrame(root, text="Select Result File")
        file_frame.place(height=100, width=400, rely=0.65, relx=0)
        # Buttons
        button1 = tk.Button(file_frame, text="Browse A File", command=lambda: self.File_dialog())
        button1.place(rely=0.65, relx=0.50)

        ## Treeview Widget
        self.tv1 = ttk.Treeview(frame1)
        self.tv1.place(relheight=1, relwidth=1)

        treescrolly = tk.Scrollbar(frame1, orient="vertical", command=self.tv1.yview)
        treescrollx = tk.Scrollbar(frame1, orient="horizontal", command=self.tv1.xview)
        self.tv1.configure(xscrollcommand=treescrollx.set,
                           yscrollcommand=treescrolly.set)  # assign the scrollbars to the Treeview Widget
        treescrollx.pack(side="bottom", fill="x")
        treescrolly.pack(side="right", fill="y")
        root.mainloop()

    def File_dialog(self):
        filename = filedialog.askopenfilename(title="Select A File",
                                              filetype=(("xlsx files", "*.xlsx"), ("All Files", "*.*")))
        self.Load_excel_data(filename)
        return None

    def Load_excel_data(self, filename):

        file_path = filename
        try:
            excel_filename = r"{}".format(file_path)
            if excel_filename[-4:] == ".csv":
                df = pd.read_csv(excel_filename)
            else:
                df = pd.read_excel(excel_filename)

        except ValueError:
            tk.messagebox.showerror("Information", "The file you have chosen is invalid")
            return None
        except FileNotFoundError:
            tk.messagebox.showerror("Information", f"No such file as {file_path}")
            return None

        self.tv1["column"] = list(df.columns)
        self.tv1["show"] = "headings"
        for column in self.tv1["columns"]:
            self.tv1.heading(column, text=column)

        data_frame_rows = df.to_numpy().tolist()
        for row in data_frame_rows:
            self.tv1.insert("", "end", values=row)
        return None
